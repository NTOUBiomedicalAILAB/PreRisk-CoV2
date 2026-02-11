#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
PreRisk-CoV2: Pre-exposure Risk Assessment for SARS-CoV-2
==========================================================
A unified machine learning framework for predicting SARS-CoV-2 susceptibility
using Serum protein biomarkers with KNN-GA approach.

Author: NTOU Biomedical AI LAB
GitHub: https://github.com/NTOUBiomedicalAILAB/PreRisk-CoV2
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import argparse
import os
import sys
import time
import random
from math import sqrt
from sklearn import preprocessing
from sklearn.metrics import (confusion_matrix, roc_curve, roc_auc_score,
                            precision_recall_curve, average_precision_score)
from sklearn.model_selection import LeaveOneOut
from sklearn.neighbors import KNeighborsClassifier
from imblearn.over_sampling import SMOTE
import openpyxl
import warnings
warnings.filterwarnings('ignore')

###############################################################################
# UTILITY FUNCTIONS
###############################################################################

def print_banner():
    """Print PreRisk-CoV2 banner."""
    banner = """
    ╔═══════════════════════════════════════════════════════════════╗
    ║                      PreRisk-CoV2                             ║
    ║         SARS-CoV-2 Pre-exposure Risk Assessment               ║
    ║              KNN-GA Protein Biomarker Framework               ║
    ╚═══════════════════════════════════════════════════════════════╝
    """
    print(banner)

def missing_counts(data):
    """Calculate missing value statistics for DataFrame columns.

    Args:
        data (pd.DataFrame): Input dataframe

    Returns:
        pd.DataFrame: Missing count statistics
    """
    missing = data.isnull().sum()
    missing = missing[missing > 0]
    missing.sort_values(inplace=True)
    missing_count = pd.DataFrame({
        'ColumnName': missing.index,
        'MissingCount': missing.values
    })
    missing_count['Percentage(%)'] = missing_count['MissingCount'].apply(
        lambda x: round(x / data.shape[0] * 100, 2)
    )
    return missing_count

def data_processing(df):
    """Process input CSV data: label encoding, feature extraction, normalization.

    Args:
        df (pd.DataFrame): Input dataframe with columns ['sample ID', 'PCR result', proteins...]

    Returns:
        tuple: (SampleID, ProteinID, Features, Label)
    """
    sample_id = df['sample ID'].values
    protein_id = df.columns.tolist()[2:94]  # 92 proteins

    df_processed = df.drop(['sample ID'], axis=1)
    df_processed['PCR result'] = df_processed['PCR result'].map({
        'Not': 0, 'Detected': 1
    }).astype(int)

    ndarray = df_processed.values
    label = ndarray[:, 0]
    features = ndarray[:, 1:]

    # MinMax normalization
    scaler = preprocessing.MinMaxScaler(feature_range=(0, 1))
    features = scaler.fit_transform(features)

    return sample_id, protein_id, features, label

def compute_metrics(y_true, y_pred, y_proba):
    """Compute comprehensive classification metrics.

    Args:
        y_true (np.array): True labels
        y_pred (np.array): Predicted labels
        y_proba (np.array): Prediction probabilities

    Returns:
        dict: Dictionary of performance metrics
    """
    cm = confusion_matrix(y_true, y_pred)
    tn, fp, fn, tp = cm.ravel()

    accuracy = (tp + tn) / (tp + fp + fn + tn)
    sensitivity = tp / (tp + fn) if (tp + fn) > 0 else 0
    specificity = tn / (tn + fp) if (tn + fp) > 0 else 0
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0
    f1_score = 2 * precision * sensitivity / (precision + sensitivity) if (precision + sensitivity) > 0 else 0

    auroc = roc_auc_score(y_true, y_proba)
    auprc = average_precision_score(y_true, y_proba)

    mcc_denom = sqrt((tp + fp) * (tp + fn) * (tn + fp) * (tn + fn))
    mcc = (tp * tn - fp * fn) / mcc_denom if mcc_denom > 0 else 0

    return {
        'accuracy': accuracy,
        'sensitivity': sensitivity,
        'specificity': specificity,
        'precision': precision,
        'f1_score': f1_score,
        'auroc': auroc,
        'auprc': auprc,
        'mcc': mcc
    }

def plot_roc_pr_curves(y_true, y_proba, save_path=None):
    """Generate ROC and PR curves.

    Args:
        y_true (np.array): True labels
        y_proba (np.array): Prediction probabilities
        save_path (str, optional): Path to save figure
    """
    fig = plt.figure(figsize=(16, 6))

    # ROC Curve
    plt.subplot(121)
    fpr, tpr, _ = roc_curve(y_true, y_proba)
    auroc = roc_auc_score(y_true, y_proba)
    plt.plot(fpr, tpr, marker='.', label=f'AUROC = {auroc:.3f}')
    plt.plot([0, 1], [0, 1], 'r--')
    plt.fill_between(fpr, tpr, alpha=0.2)
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.0])
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')
    plt.title('ROC Curve')
    plt.legend(loc='lower right')
    plt.gca().set_aspect('equal')

    # PR Curve
    plt.subplot(122)
    precision, recall, _ = precision_recall_curve(y_true, y_proba)
    auprc = average_precision_score(y_true, y_proba)
    plt.plot(recall, precision, marker='.', label=f'AUPRC = {auprc:.3f}')
    plt.fill_between(recall, precision, alpha=0.2)
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.0])
    plt.xlabel('Recall')
    plt.ylabel('Precision')
    plt.title('Precision-Recall Curve')
    plt.legend(loc='lower right')
    plt.gca().set_aspect('equal')

    plt.tight_layout()
    if save_path:
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.show()

def build_knn_model(n_neighbors=5, leaf_size=30, algorithm='auto', weights='uniform', p=2):
    """Build KNN classifier with specified hyperparameters.

    Args:
        n_neighbors (int): Number of neighbors
        leaf_size (int): Leaf size for tree algorithms
        algorithm (str): Algorithm type ('auto', 'ball_tree', 'kd_tree', 'brute')
        weights (str): Weight function ('uniform', 'distance')
        p (int): Power parameter for Minkowski metric

    Returns:
        KNeighborsClassifier: Configured KNN model
    """
    return KNeighborsClassifier(
        n_neighbors=n_neighbors,
        leaf_size=leaf_size,
        algorithm=algorithm,
        weights=weights,
        p=p
    )

def save_results_to_excel(results, output_path, sheet_name, mode='internal'):
    """Save validation results to Excel file.

    Args:
        results (np.array): Results array with metrics
        output_path (str): Excel file path
        sheet_name (str): Sheet name
        mode (str): 'internal' or 'external' validation mode
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    try:
        workbook = openpyxl.load_workbook(output_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)

    workbook.create_sheet(sheet_name, 0)
    worksheet = workbook.active

    n_iter = len(results)

    # Header
    if mode == 'internal':
        worksheet.append(['以下為每輪交叉驗證之結果:'])
    else:
        worksheet.append(['以下為每輪預測之結果:'])

    worksheet.append([''])
    worksheet.append(['Iteration', 'Accuracy', 'Sensitivity', 'Specificity', 
                     'Precision', 'F1-Score', 'AUROC', 'AUPRC', 'MCC'])

    # Data rows
    for row in results:
        worksheet.append(row.tolist())

    # Statistics
    worksheet.append([''])
    worksheet.append(['', 'Accuracy', 'Sensitivity', 'Specificity', 
                     'Precision', 'F1-Score', 'AUROC', 'AUPRC', 'MCC'])

    means = np.mean(results[:, 1:], axis=0)
    stds = np.std(results[:, 1:], axis=0)

    worksheet.append(['Mean'] + means.tolist())
    worksheet.append(['Std Dev'] + stds.tolist())

    formatted = ['Formatted']
    for mean, std in zip(means, stds):
        if mean > 1:  # Percentage metrics
            formatted.append(f'{mean*100:.2f} ± {std*100:.2f}%')
        else:  # Decimal metrics
            formatted.append(f'{mean:.4f} ± {std:.4f}')

    worksheet.append(formatted)

    workbook.save(output_path)
    print(f"[INFO] Results saved to: {output_path}")

###############################################################################
# VALIDATION MODES
###############################################################################

def internal_validation(args):
    """Perform internal validation using Leave-One-Out Cross-Validation (LOOCV).

    Args:
        args: Command-line arguments from argparse
    """
    print("\n" + "="*70)
    print("INTERNAL VALIDATION MODE (LOOCV)")
    print("="*70)

    # Load data
    print(f"[INFO] Loading data from: {args.input}")
    df = pd.read_csv(args.input)
    sample_id, protein_id, features, label = data_processing(df)

    print(f"[INFO] Dataset shape: {features.shape}")
    print(f"[INFO] Class distribution: {np.bincount(label.astype(int))}")

    # Select protein biomarkers
    if args.protein_indices:
        protein_indices = args.protein_indices
    else:
        protein_indices = [3, 50, 40, 36, 83]  # Default 5-protein panel

    print(f"[INFO] Selected proteins: {[protein_id[i] for i in protein_indices]}")
    features_selected = features[:, protein_indices]

    # LOOCV setup
    loo = LeaveOneOut()
    n_samples = len(features_selected)
    results = np.zeros((args.n_iterations, 9))

    print(f"[INFO] Running {args.n_iterations} iterations of LOOCV...")
    start_time = time.time()

    for iteration in range(args.n_iterations):
        y_true_all = []
        y_pred_all = []
        y_proba_all = []

        for train_idx, test_idx in loo.split(features_selected):
            X_train, X_test = features_selected[train_idx], features_selected[test_idx]
            y_train, y_test = label[train_idx], label[test_idx]

            # SMOTE oversampling
            if args.use_smote and len(np.unique(y_train)) > 1:
                try:
                    smote = SMOTE(k_neighbors=min(5, sum(y_train == 0) - 1, sum(y_train == 1) - 1))
                    X_train, y_train = smote.fit_resample(X_train, y_train)
                except:
                    pass

            # Train KNN
            model = build_knn_model(
                n_neighbors=args.n_neighbors,
                leaf_size=args.leaf_size,
                algorithm=args.algorithm,
                weights=args.weights,
                p=args.p
            )
            model.fit(X_train, y_train)

            # Predict
            y_pred = model.predict(X_test)
            y_proba = model.predict_proba(X_test)[:, 1]

            y_true_all.append(y_test[0])
            y_pred_all.append(y_pred[0])
            y_proba_all.append(y_proba[0])

        # Compute metrics
        metrics = compute_metrics(
            np.array(y_true_all),
            np.array(y_pred_all),
            np.array(y_proba_all)
        )

        results[iteration, 0] = iteration + 1
        results[iteration, 1] = metrics['accuracy']
        results[iteration, 2] = metrics['sensitivity']
        results[iteration, 3] = metrics['specificity']
        results[iteration, 4] = metrics['precision']
        results[iteration, 5] = metrics['f1_score']
        results[iteration, 6] = metrics['auroc']
        results[iteration, 7] = metrics['auprc']
        results[iteration, 8] = metrics['mcc']

        if args.verbose:
            print(f"  Iter {iteration+1}/{args.n_iterations}: "
                  f"Acc={metrics['accuracy']:.3f}, "
                  f"AUROC={metrics['auroc']:.3f}, "
                  f"AUPRC={metrics['auprc']:.3f}")

        # Plot curves for last iteration
        if args.plot_curves and iteration == args.n_iterations - 1:
            plot_roc_pr_curves(
                np.array(y_true_all),
                np.array(y_proba_all),
                save_path=os.path.join(args.output_dir, 'internal_roc_pr.png')
            )

    elapsed = time.time() - start_time
    print(f"[INFO] Total time: {elapsed:.2f} seconds")

    # Print summary
    print("\n" + "-"*70)
    print("SUMMARY STATISTICS")
    print("-"*70)
    means = np.mean(results[:, 1:], axis=0)
    stds = np.std(results[:, 1:], axis=0)
    metric_names = ['Accuracy', 'Sensitivity', 'Specificity', 'Precision', 
                   'F1-Score', 'AUROC', 'AUPRC', 'MCC']

    for name, mean, std in zip(metric_names, means, stds):
        if mean > 1:
            print(f"{name:15s}: {mean*100:6.2f} ± {std*100:5.2f} %")
        else:
            print(f"{name:15s}: {mean:6.4f} ± {std:6.4f}")

    # Save results
    timestamp = time.strftime("%Y%m%d_%H%M")
    output_path = os.path.join(args.output_dir, f'internal_validation_{timestamp}.xlsx')
    save_results_to_excel(results, output_path, 'LOOCV_Results', mode='internal')

def external_validation(args):
    """Perform external validation on independent test cohort.

    Args:
        args: Command-line arguments from argparse
    """
    print("\n" + "="*70)
    print("EXTERNAL VALIDATION MODE")
    print("="*70)

    # Load data
    print(f"[INFO] Loading training data from: {args.train_input}")
    train_df = pd.read_csv(args.train_input)
    _, protein_id_train, train_features, train_label = data_processing(train_df)

    print(f"[INFO] Loading test data from: {args.test_input}")
    test_df = pd.read_csv(args.test_input)
    _, protein_id_test, test_features, test_label = data_processing(test_df)

    print(f"[INFO] Training set shape: {train_features.shape}")
    print(f"[INFO] Test set shape: {test_features.shape}")
    print(f"[INFO] Training class distribution: {np.bincount(train_label.astype(int))}")
    print(f"[INFO] Test class distribution: {np.bincount(test_label.astype(int))}")

    # Select protein biomarkers
    if args.protein_indices:
        protein_indices = args.protein_indices
    else:
        protein_indices = [3, 50, 40, 36, 83]

    print(f"[INFO] Selected proteins: {[protein_id_train[i] for i in protein_indices]}")
    train_features_selected = train_features[:, protein_indices]
    test_features_selected = test_features[:, protein_indices]

    results = np.zeros((args.n_iterations, 9))

    print(f"[INFO] Running {args.n_iterations} iterations...")
    start_time = time.time()

    for iteration in range(args.n_iterations):
        X_train, y_train = train_features_selected, train_label

        # SMOTE oversampling
        if args.use_smote:
            smote = SMOTE(k_neighbors=5)
            X_train, y_train = smote.fit_resample(X_train, y_train)

        # Train KNN
        model = build_knn_model(
            n_neighbors=args.n_neighbors,
            leaf_size=args.leaf_size,
            algorithm=args.algorithm,
            weights=args.weights,
            p=args.p
        )
        model.fit(X_train, y_train)

        # Predict on test set
        y_pred = model.predict(test_features_selected)
        y_proba = model.predict_proba(test_features_selected)[:, 1]

        # Compute metrics
        metrics = compute_metrics(test_label, y_pred, y_proba)

        results[iteration, 0] = iteration + 1
        results[iteration, 1] = metrics['accuracy']
        results[iteration, 2] = metrics['sensitivity']
        results[iteration, 3] = metrics['specificity']
        results[iteration, 4] = metrics['precision']
        results[iteration, 5] = metrics['f1_score']
        results[iteration, 6] = metrics['auroc']
        results[iteration, 7] = metrics['auprc']
        results[iteration, 8] = metrics['mcc']

        if args.verbose:
            print(f"  Iter {iteration+1}/{args.n_iterations}: "
                  f"Acc={metrics['accuracy']:.3f}, "
                  f"AUROC={metrics['auroc']:.3f}, "
                  f"AUPRC={metrics['auprc']:.3f}")

        # Plot curves for last iteration
        if args.plot_curves and iteration == args.n_iterations - 1:
            plot_roc_pr_curves(
                test_label,
                y_proba,
                save_path=os.path.join(args.output_dir, 'external_roc_pr.png')
            )

    elapsed = time.time() - start_time
    print(f"[INFO] Total time: {elapsed:.2f} seconds")

    # Print summary
    print("\n" + "-"*70)
    print("SUMMARY STATISTICS")
    print("-"*70)
    means = np.mean(results[:, 1:], axis=0)
    stds = np.std(results[:, 1:], axis=0)
    metric_names = ['Accuracy', 'Sensitivity', 'Specificity', 'Precision', 
                   'F1-Score', 'AUROC', 'AUPRC', 'MCC']

    for name, mean, std in zip(metric_names, means, stds):
        if mean > 1:
            print(f"{name:15s}: {mean*100:6.2f} ± {std*100:5.2f} %")
        else:
            print(f"{name:15s}: {mean:6.4f} ± {std:6.4f}")

    # Save results
    timestamp = time.strftime("%Y%m%d_%H%M")
    output_path = os.path.join(args.output_dir, f'external_validation_{timestamp}.xlsx')
    save_results_to_excel(results, output_path, 'External_Results', mode='external')

###############################################################################
# MAIN
###############################################################################

def main():
    """Main entry point for PreRisk-CoV2."""
    parser = argparse.ArgumentParser(
        description='PreRisk-CoV2: SARS-CoV-2 Pre-exposure Risk Assessment',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Internal validation (LOOCV)
  python prerisk_cov2.py --mode internal --input Discovery.csv --n-iterations 100

  # External validation
  python prerisk_cov2.py --mode external --train-input Discovery.csv --test-input Validation.csv

  # With custom KNN parameters
  python prerisk_cov2.py --mode internal --input Discovery.csv --n-neighbors 7 --weights distance

  # With visualization
  python prerisk_cov2.py --mode internal --input Discovery.csv --plot-curves --verbose
        """
    )

    # Mode selection
    parser.add_argument('--mode', type=str, required=True, choices=['internal', 'external'],
                       help='Validation mode: internal (LOOCV) or external (independent cohort)')

    # Data inputs
    parser.add_argument('--input', type=str, 
                       help='Input CSV file for internal validation')
    parser.add_argument('--train-input', type=str, 
                       help='Training CSV file for external validation')
    parser.add_argument('--test-input', type=str, 
                       help='Test CSV file for external validation')

    # Feature selection
    parser.add_argument('--protein-indices', type=int, nargs='+', default=None,
                       help='Protein column indices (0-based). Default: [3,50,40,36,83]')

    # KNN hyperparameters
    parser.add_argument('--n-neighbors', type=int, default=5,
                       help='Number of neighbors for KNN (default: 5)')
    parser.add_argument('--leaf-size', type=int, default=30,
                       help='Leaf size for tree-based algorithms (default: 30)')
    parser.add_argument('--algorithm', type=str, default='auto',
                       choices=['auto', 'ball_tree', 'kd_tree', 'brute'],
                       help='KNN algorithm (default: auto)')
    parser.add_argument('--weights', type=str, default='uniform',
                       choices=['uniform', 'distance'],
                       help='Weight function (default: uniform)')
    parser.add_argument('--p', type=int, default=2,
                       help='Power parameter for Minkowski metric (default: 2)')

    # Training options
    parser.add_argument('--use-smote', action='store_true', default=False,
                       help='Enable SMOTE oversampling for class imbalance')
    parser.add_argument('--n-iterations', type=int, default=100,
                       help='Number of validation iterations (default: 100)')

    # Output options
    parser.add_argument('--output-dir', type=str, default='./results',
                       help='Output directory for results (default: ./results)')
    parser.add_argument('--plot-curves', action='store_true', default=False,
                       help='Generate ROC and PR curves')
    parser.add_argument('--verbose', action='store_true', default=False,
                       help='Print detailed progress')

    args = parser.parse_args()

    # Validate arguments
    if args.mode == 'internal' and not args.input:
        parser.error("--input is required for internal validation")
    if args.mode == 'external' and (not args.train_input or not args.test_input):
        parser.error("--train-input and --test-input are required for external validation")

    # Print banner
    print_banner()

    # Create output directory
    os.makedirs(args.output_dir, exist_ok=True)

    # Run validation
    if args.mode == 'internal':
        internal_validation(args)
    else:
        external_validation(args)

    print("\n" + "="*70)
    print("PreRisk-CoV2 completed successfully!")
    print("="*70 + "\n")

if __name__ == '__main__':
    main()
