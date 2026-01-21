###################################################   導入套件   ###########################################################
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt  # 繪圖相關套件

from sklearn import preprocessing
from sklearn.preprocessing import OneHotEncoder  # 獨熱編碼(OneHot)
import os
import openpyxl
import time
import itertools
import random

from keras.models import Sequential  # 線性堆疊模型
from keras.layers import Dense  #
from keras.optimizers import SGD
from keras.layers import Dropout, LeakyReLU, ELU, BatchNormalization, Activation
from keras import backend as K
from keras.layers import LeakyReLU
from sklearn.metrics import confusion_matrix
from sklearn.metrics import confusion_matrix, roc_curve, roc_auc_score
from sklearn.metrics import precision_recall_curve, average_precision_score
from keras.optimizers import SGD, Adagrad, Adadelta, RMSprop
from sklearn.metrics import log_loss
from sklearn.model_selection import LeaveOneOut
from collections import Counter
from imblearn.over_sampling import SMOTE
from imblearn.under_sampling import TomekLinks
from keras.models import load_model
from math import *
from sklearn.feature_selection import mutual_info_classif, SelectKBest
from statistics import *

from sklearn.ensemble import RandomForestClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.svm import SVC
from sklearn.tree import DecisionTreeClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.naive_bayes import GaussianNB, MultinomialNB
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import AdaBoostClassifier, GradientBoostingClassifier
from xgboost import XGBClassifier
from sklearn.model_selection import LeaveOneOut


start_time = time.time()  # 程式開始時間擷取



#################################################### 函數 ###############################################################




# 定義用來統計欄位缺漏值總數的函數
def Missing_Counts(Data):
    missing = Data.isnull().sum()  # 計算欄位中缺漏值的數量
    missing = missing[missing > 0]
    missing.sort_values(inplace=True)

    Missing_Count = pd.DataFrame(
        {'ColumnName': missing.index, 'MissingCount': missing.values})  # Convert Series to DataFrame
    Missing_Count['Percentage(%)'] = Missing_Count['MissingCount'].apply(lambda x: round(x / Data.shape[0] * 100, 2))
    return Missing_Count



def data_processing(all_df):

    SampleID = all_df['sample ID']

    SampleID = SampleID.T

    ProteinID = all_df.columns.tolist()

    ProteinID = ProteinID[2:]

    ProteinID = ProteinID[:92]

    #print(ProteinID)

    # #刪除sample ID

    df = all_df.drop(['sample ID'], axis=1)
    # print(df[df.isnull().values==True])  #找出有nan值的那列資料

    # 確診為1 沒有確診為0(必要)
    df['PCR result'] = df['PCR result'].map({'Not': 0, 'Detected': 1}).astype(int)

    ndarray = df.values

    # display(df)
    # 查詢具空值的欄位
    # print( 'train :' )
    # display( Missing_Counts(df) )

    # 取feature 與label
    Label = ndarray[:, 0]
    Features = ndarray[:, 1:]


    # 將特徵值轉換成0~1之間
    minmax_scale = preprocessing.MinMaxScaler(feature_range=(0, 1))
    Features = minmax_scale.fit_transform(Features)


    return SampleID, ndarray, Features, Label, ProteinID


def PR_ROC_curve(test_label, dnn_probs):
    fig = plt.gcf()
    fig.set_size_inches(16, 6)
    plt_1 = plt.subplot(121)
    plt_1.set_box_aspect(1)
    dnn_auc = roc_auc_score(test_label, dnn_probs)
    dnn_fpr, dnn_tpr, _ = roc_curve(test_label, dnn_probs)

    plt.plot(dnn_fpr, dnn_tpr, marker='.', label=' (AUROC = %0.3f)' % dnn_auc)

    plt.plot([0, 1], [0, 1], color='red', linestyle='--')  # lw 線寬
    plt.title('ROC')
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')

    plt.fill_between(dnn_fpr, dnn_tpr, color='gray', alpha=0.2)

    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.0])
    plt.legend(loc='lower right')


    plt_2 = plt.subplot(122)
    plt_2.set_box_aspect(1)
    precision, recall, thresholds = precision_recall_curve(test_label, dnn_probs)
    dnn_aup = average_precision_score(test_label, dnn_probs)

    plt.plot(recall, precision, label=' (AUPRC = %0.3f)' % dnn_aup)

    fontsize = 14
    plt.title("PR")  # 標題
    plt.xlabel("Recall", fontsize=fontsize)
    plt.ylabel("Precision", fontsize=fontsize)

    plt.fill_between(recall, precision, color='gray', alpha=0.2)


    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.0])
    plt.legend(loc='lower right')

    print("========================")
    for i in range(len(test_label)) : print(f"{test_label[i]}  {dnn_probs[i]}")

    plt.show()


def excel_save(data_save, sheet_name, excel_name,location):
    global SampleID,test_save

    #輸出結果至excel
    os.chdir(location)  # Colab路徑


    import openpyxl

    try:
        workbook = openpyxl.load_workbook(excel_name)
        worksheet = workbook.active  # 取得作用中的工作表
    except FileNotFoundError:
        workbook = openpyxl.Workbook() # 新工作表

    print(sheet_name)

    #開啟excel
    workbook.create_sheet(sheet_name, 0)  # 新工作表 指定索引位置
    worksheet = workbook.active # 取得作用中的工作表
    worksheet.append(['以下為每輪交叉驗證之結果:'])
    worksheet.append([''])

    worksheet.append([' ', 'Accuracy', 'Specificity', 'Sensitivity', 'Precision', "AUROC", 'AUPRC', 'MCC', 'F1_score'])

    for i in data_save :
        worksheet.append(i.tolist())

    worksheet.append([''])
    worksheet.append(['','accuracy', 'Specificity', 'Sensitivity', 'Precision', "AUROC", 'AUPRC', 'F1_score', 'MCC'])

    worksheet.append(['總平均', sum(data_save[:,1]) / loop, sum(data_save[:,2]) / loop, sum(data_save[:,3]) / loop, sum(data_save[:,4]) / loop, sum(data_save[:,5]) / loop, sum(data_save[:,6]) / loop, sum(data_save[:,8]) / loop, sum(data_save[:,7]) / loop])

    worksheet.append(['標準差', np.std(data_save[:, 1]), np.std(data_save[:, 2]), np.std(data_save[:, 3]), np.std(data_save[:, 4]), np.std(data_save[:, 5]), np.std(data_save[:, 6]), np.std(data_save[:, 8]), np.std(data_save[:, 7])])

    worksheet.append(
        ['數據', str(round(sum(data_save[:, 1]) / loop * 100, 2)) + ' ± ' + str(round(np.std(data_save[:, 1]) * 100, 2))
            , str(round(sum(data_save[:, 2]) / loop * 100, 2)) + ' ± ' + str(round(np.std(data_save[:, 2]) * 100, 2))
            , str(round(sum(data_save[:, 3]) / loop * 100, 2)) + ' ± ' + str(round(np.std(data_save[:, 3]) * 100, 2))
            , str(round(sum(data_save[:, 4]) / loop * 100, 2)) + ' ± ' + str(round(np.std(data_save[:, 4]) * 100, 2))
            , str(round(sum(data_save[:, 5]) / loop, 4)) + ' ± ' + str(round(np.std(data_save[:, 5]), 4))
            , str(round(sum(data_save[:, 6]) / loop, 4)) + ' ± ' + str(round(np.std(data_save[:, 6]), 4))
            , str(round(sum(data_save[:, 8]) / loop, 4)) + ' ± ' + str(round(np.std(data_save[:, 8]), 4))
            , str(round(sum(data_save[:, 7]) / loop * 100, 2)) + ' ± ' + str(round(np.std(data_save[:, 7]) * 100, 2)) ])



    workbook.save(excel_name)


# 取得時間
localtime = time.localtime()
local_time = time.strftime("%Y%m%d_%H%M%p", localtime)


#####################################################   參數設定   #######################################################
# excel 設定
save_new = 1
file = 'Discovery.csv'

# 模型參數(常用)
train_process = 0  # 訓練過程
loop_result = 0  # 是否觀看每次交叉驗證過程
smote = 0  # {0: 沒有SMOTE, 1: SMOTE}


ID_index = [3, 50, 40, 36, 83] # 測試蛋白質編號
random_protein = 7 #隨機選擇蛋白質數量
pr_roc = 0 # 印出 PR、ROC 曲線


excel_name = '內部驗證_' + local_time + '.xlsx'  # + local_time
####################################################   匯入資料   ########################################################

loop = 100
excel_location = '\\輸出結果'


if save_new == 1:
    new = 1  # 是否要創建新的excel
    save = 1  # 是否存到excel  (excel記得關掉)
else:
    new = 0  # 是否要創建新的excel
    save = 0  # 是否存到excel  (excel記得關掉)


# 輸入資料途徑
excel_input = 'Discovery.csv'

all_df = pd.read_csv(excel_input)

SampleID, ndarray, Features, Label, ProteinID = data_processing(all_df)


###################################################   主程式   #########################################################

if train_process == 1: loop_result = 1

if random_protein > 0 : excel_name = '內部驗證_隨機_' + str(random_protein) +'個蛋白質.xlsx'

n_neighbors = [0]
leaf_size = [11]
algo = [3]
weigh = [2]
p = [2]

data_use = [ID_index] * 100
accuracy_indice = [tuple(item) for item in data_use]
print(accuracy_indice)


loo = LeaveOneOut()

best_save = np.zeros((1000, 14), dtype=float)  # 儲存結果

result_count = 1

good = np.zeros(len(Features))



for n_neighbors,leaf_size,algo,weigh,p in zip(n_neighbors,leaf_size,algo,weigh,p) :

    data_save = np.zeros((loop, 9), dtype=float)  # 儲存結果

    total_start_time = time.time()

    for i in range(loop) :

        ans = []
        pred = []
        probs = []


        if random_protein > 0 :
            ID_index = []
            ID_index = random.sample(range(0,92),random_protein)  # 基因演算法_OAX
            data_use = [ID_index] * 100
            accuracy_indice = [tuple(item) for item in data_use]
            print(accuracy_indice)



        for loo_count, (train_index, test_index) in enumerate(loo.split(Features)) :  # 切資料

            accuracy_protein = []

            indice = accuracy_indice[loo_count]

            for j in indice : accuracy_protein.append(ProteinID[j])


            test_number = test_index[0]

            train_data, test_data = Features[train_index, :], Features[test_index, :]  # 切訓練、測試資料

            train_label, test_label = Label[train_index], Label[test_index]  # 切訓練、測試資料


            if algo == 1:
                algorithm = 'auto'
            elif algo == 2:
                algorithm = 'ball_tree'
            elif algo == 3:
                algorithm = 'kd_tree'
            elif algo == 4:
                algorithm = 'brute'

            if weigh == 1:
                weights = 'uniform'
            elif weigh == 2:
                weights = 'distance'

            if n_neighbors == 0:
                model = KNeighborsClassifier(n_neighbors=5)
            else:
                model = KNeighborsClassifier(n_neighbors=n_neighbors, leaf_size=leaf_size, algorithm=algorithm, weights=weights,p=p)


            train_data = train_data[:, np.array(indice)]
            test_data = test_data[:, np.array(indice)]


            if smote == 1:
                train_data, train_label = SMOTE(k_neighbors=5).fit_resample(train_data,train_label)


            model.fit(train_data, train_label)  # 將最好的特徵組合跟訓練資料進行模型訓練

            prediction = model.predict(test_data)  # 將最好的特徵組合跟測試資料進行模型預測

            proba = model.predict_proba(test_data)
            proba = proba[:, 1]



            ans.append(test_label[0])
            pred.append(prediction[0])
            probs.append(proba[0])

            if prediction[0] == 1 : good[loo_count] = good[loo_count] + 1



        if pr_roc == 1 :
            PR_ROC_curve(ans, probs)
            print("pred = ", pred)
            print("probs = ", probs)



        cm_sklearn = confusion_matrix(ans, pred)


        TP = cm_sklearn[1, 1]
        FN = cm_sklearn[1, 0]
        FP = cm_sklearn[0, 1]
        TN = cm_sklearn[0, 0]


        accuracy = (TP + TN) / (TP + FP + FN + TN)
        recall = TP / (TP + FN)
        precision = TP / (TP + FP)
        F1_score = 2 / ((1 / precision) + (1 / recall))
        AUPRC = average_precision_score(ans, probs)
        AUROC = roc_auc_score(ans, probs)
        specificity = TN / (TN + FP)
        MCC = (TP * TN - FP * FN) / sqrt((TP + FP) * (TP + FN) * (TN + FP) * (TN + FN))

        data_save[i, 0] = i + 1
        data_save[i, 1] = accuracy
        data_save[i, 2] = specificity
        data_save[i, 3] = recall
        data_save[i, 4] = precision
        data_save[i, 5] = AUROC
        data_save[i, 6] = AUPRC
        data_save[i, 7] = MCC
        data_save[i, 8] = F1_score

        if loop_result == 1 :
            print('============================================最終結果=====================================================')
            print('第',i+1,'次預測')
            print('accuracy =', round(accuracy * 100, 2), '%')
            print('specificity =',round(specificity * 100, 2), '%')
            print('recall =', round(recall * 100, 2), '%')
            print('precision =', round(precision * 100, 2), '%')
            print('AUROC =', round(AUROC,4))
            print('AUPRC =', round(AUPRC,4))
            print('F1_score =', round(F1_score * 100, 2), '%')
            print('MCC =', round(MCC, 4))

            print("test_label  : ", ans)
            print("prediction  : ", pred)
            print(" ")



    if save == 1:
        sheet_name = '編號 ' + str(result_count)  # 'layer ' + str(layer) + ' batch ' + str(batch_size) + ' units ' + str(units) +
        excel_save(data_save, sheet_name, excel_name, excel_location)  # 輸出結果至excel


    print('===============================================================================================')
    print('n_neighbors', n_neighbors, ' leaf_size ', leaf_size, ' algorithm ', algo, 'weight', weigh, 'p', p)
    print('(上面為模型隱藏層資訊)')
    print(loop, '次交叉驗證平均後:')
    print('accuracy    = ', round(sum(data_save[:, 1]) / loop * 100, 2), '±', round(np.std(data_save[:, 1]) * 100, 2))
    print('specificity = ', round(sum(data_save[:, 2]) / loop * 100, 2), '±', round(np.std(data_save[:, 2]) * 100, 2))
    print('recall      = ', round(sum(data_save[:, 3]) / loop * 100, 2), '±', round(np.std(data_save[:, 3]) * 100, 2))
    print('precision   = ', round(sum(data_save[:, 4]) / loop * 100, 2), '±', round(np.std(data_save[:, 4]) * 100, 2))
    print('AUROC       = ', round(sum(data_save[:, 5]) / loop, 4), '±', round(np.std(data_save[:, 5]), 4))
    print('AUPRC       = ', round(sum(data_save[:, 6]) / loop, 4), '±', round(np.std(data_save[:, 6]), 4))
    print('MCC         = ', round(sum(data_save[:, 7]) / loop, 4), '±', round(np.std(data_save[:, 7]), 4))
    print('F1_score    = ', round(sum(data_save[:, 8]) / loop * 100, 2), '±', round(np.std(data_save[:, 8]) * 100, 2))

    print(" ")

    if sum(data_save[:, 1]) / loop > 0.7 :

        best_save[result_count, 0] = result_count
        best_save[result_count, 1] = sum(data_save[:, 1]) / loop
        best_save[result_count, 2] = sum(data_save[:, 2]) / loop
        best_save[result_count, 3] = sum(data_save[:, 3]) / loop
        best_save[result_count, 4] = sum(data_save[:, 4]) / loop
        best_save[result_count, 5] = sum(data_save[:, 6]) / loop
        best_save[result_count, 6] = sum(data_save[:, 5]) / loop
        best_save[result_count, 7] = sum(data_save[:, 8]) / loop
        best_save[result_count, 8] = sum(data_save[:, 7]) / loop
        best_save[result_count, 9] = n_neighbors
        best_save[result_count, 10] = leaf_size
        best_save[result_count, 11] = algo
        best_save[result_count, 12] = weigh
        best_save[result_count, 13] = p

        result_count = result_count + 1

        if result_count > loop : break

    total_end_time = time.time()  # 程式結束時間擷取
    print("總用時 : ", total_end_time - total_start_time, ' s')  # 總用時




end_time = time.time()  # 程式結束時間擷取


print('===============================================================================================')
print(" ")
print("程式總用時 : ", end_time - start_time, ' s')  # 總用時
print(" ")