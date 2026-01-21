###################################################   導入套件   ###########################################################
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt  # 繪圖相關套件
from sklearn import preprocessing
from sklearn.preprocessing import OneHotEncoder  # 獨熱編碼(OneHot)
import os
import openpyxl
import time
import itertools

from keras.models import Sequential  # 線性堆疊模型
from keras.layers import Dense  #
from keras.optimizers import SGD
from keras.layers import Dropout, LeakyReLU, ELU, BatchNormalization, Activation
from keras import backend as K
from keras.layers import LeakyReLU
from sklearn.metrics import confusion_matrix
from sklearn.metrics import confusion_matrix, roc_curve, roc_auc_score
from sklearn.metrics import precision_recall_curve, average_precision_score
from keras.optimizers import SGD, Adagrad, Adadelta, RMSprop
from sklearn.metrics import log_loss
from sklearn.model_selection import LeaveOneOut
from collections import Counter
from imblearn.over_sampling import SMOTE, BorderlineSMOTE, ADASYN, SVMSMOTE, KMeansSMOTE, SMOTENC, RandomOverSampler
from imblearn.under_sampling import TomekLinks
from keras.models import load_model
from math import *
from sklearn.feature_selection import mutual_info_classif, SelectKBest
from statistics import *

from sklearn.ensemble import RandomForestClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.svm import SVC
from sklearn.tree import DecisionTreeClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.naive_bayes import GaussianNB, MultinomialNB
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import LeaveOneOut

###################################################   函式   ###########################################################

# 定義用來統計欄位缺漏值總數的函數
def Missing_Counts(Data):
    missing = Data.isnull().sum()  # 計算欄位中缺漏值的數量
    missing = missing[missing > 0]
    missing.sort_values(inplace=True)

    Missing_Count = pd.DataFrame(
        {'ColumnName': missing.index, 'MissingCount': missing.values})  # Convert Series to DataFrame
    Missing_Count['Percentage(%)'] = Missing_Count['MissingCount'].apply(lambda x: round(x / Data.shape[0] * 100, 2))
    return Missing_Count


def score(ans, prediction, probs) :
    global ProteinID, accuracy_indice, goodbad

    i = loop_count

    cm_sklearn = confusion_matrix(ans, prediction)

    TP = cm_sklearn[1, 1]
    FP = cm_sklearn[0, 1]
    FN = cm_sklearn[1, 0]
    TN = cm_sklearn[0, 0]

    accuracy = (TP + TN) / (TP + FP + FN + TN)
    Sensitivity = TP / (TP + FN)
    precision = TP / (TP + FP)
    F1_score = 2 / ((1 / precision) + (1 / Sensitivity))
    AUPRC = average_precision_score(ans, probs)
    AUROC = roc_auc_score(ans, probs)
    Specificity = TN / (TN + FP)
    MCC = (TP * TN - FP * FN) / sqrt((TP + FP) * (TP + FN) * (TN + FP) * (TN + FN))


    data_save[i, 0] = i + 1
    data_save[i, 1] = accuracy
    data_save[i, 2] = Sensitivity
    data_save[i, 3] = precision
    data_save[i, 4] = F1_score
    data_save[i, 5] = AUROC
    data_save[i, 6] = AUPRC
    data_save[i, 7] = MCC
    data_save[i, 8] = Specificity

    if prediction_process == 1 :
        print('===============================================================================================')
        print("特徵子集 : ", )
        for z in accuracy_indice: print(ProteinID[z])
        print("第",loop_count+1,"次預測")
        print('accuracy =', accuracy)
        print('Specificity =', Specificity)
        print('Sensitivity =', Sensitivity)
        print('precision =', precision)
        print('AUROC =', AUROC)
        print('AUPRC =', AUPRC)
        print('F1_score =', F1_score)
        print('MCC =', MCC)

        print("ans        :", ans)
        print("prediction :", prediction)

    for i in range(len(goodbad)) :
        if ans[i] == prediction[i] : goodbad[i] = goodbad[i] + 1



def data_processing(all_df):
    global SampleID, ProteinID

    # print( f'train {all_df.shape}' )
    # display( all_df.head() )

    SampleID = all_df['sample ID']

    SampleID = SampleID.T

    ProteinID = all_df.columns.tolist()

    ProteinID = ProteinID[2:]

    ProteinID = ProteinID[:92]

    #print(ProteinID)

    # #刪除sample ID

    df = all_df.drop(['sample ID'], axis=1)
    # print(df[df.isnull().values==True])  #找出有nan值的那列資料

    # 確診為1 沒有確診為0(必要)
    df['PCR result'] = df['PCR result'].map({'Not': 0, 'Detected': 1}).astype(int)

    ndarray = df.values


    # display(df)
    # 查詢具空值的欄位
    # print( 'train :' )
    # display( Missing_Counts(df) )

    # 取feature 與label
    Label = ndarray[:, 0]
    Features = ndarray[:, 1:]

    # 將特徵值轉換成0~1之間
    minmax_scale = preprocessing.MinMaxScaler(feature_range=(0, 1))
    Features = minmax_scale.fit_transform(Features)


    return Features, Label


def BestSave(data_save, result_count):
    global best_save
    best_save[result_count, 0] = result_count + 1
    best_save[result_count, 1] = sum(data_save[:, 1]) / loop
    best_save[result_count, 2] = sum(data_save[:, 4]) / loop
    best_save[result_count, 3] = sum(data_save[:, 3]) / loop
    best_save[result_count, 4] = sum(data_save[:, 7]) / loop
    best_save[result_count, 5] = sum(data_save[:, 6]) / loop
    best_save[result_count, 6] = sum(data_save[:, 5]) / loop
    best_save[result_count, 7] = sum(data_save[:, 2]) / loop
    best_save[result_count, 8] = sum(data_save[:, 8]) / loop
    best_save[result_count, 9] = n_neighbors
    best_save[result_count, 10] = leaf_size
    best_save[result_count, 11] = algo
    best_save[result_count, 12] = weigh
    best_save[result_count, 13] = p


def excel_save_best(sheet_name, excel_name, location):

    os.chdir(location)  # 鎖定路徑

    # 輸出結果至excel
    import openpyxl

    try:
        workbook = openpyxl.load_workbook(excel_name)
        worksheet = workbook.active  # 取得作用中的工作表
    except FileNotFoundError:
        workbook = openpyxl.Workbook() # 新工作表


    #開啟excel
    workbook.create_sheet(sheet_name, 0)  # 新工作表 指定索引位置
    worksheet = workbook.active # 取得作用中的工作表



    worksheet.append(['number','accuracy', 'precision', 'recall', 'F1_score', 'AUPRC', "AUROC", 'sensitivity', 'MCC', 'n_neighbors', 'leaf_size','algorithm','weight','p'])

    for i in best_save:
        worksheet.append(i.tolist())


    workbook.save(excel_name)


def excel_save(data_save, sheet_name, excel_name, location):
    global SampleID

    print(location)

    os.chdir(location)  # 鎖定路徑

    #輸出結果至excel

    try:
        workbook = openpyxl.load_workbook(excel_name)
        worksheet = workbook.active  # 取得作用中的工作表
    except FileNotFoundError:
        workbook = openpyxl.Workbook() # 新工作表

    #開啟excel
    workbook.create_sheet(sheet_name, 0)  # 新工作表 指定索引位置
    worksheet = workbook.active # 取得作用中的工作表
    # worksheet.title = sheet_name # 幫新工作表取名子

    # worksheet = workbook[sheet_name]      # 開啟工作表


    worksheet.append(['','Accuracy', 'Sensitivity', 'Precision', 'F1_score', "AUROC", 'AUPRC', 'MCC', 'Specificity'])

    worksheet.append(['總平均', sum(data_save[:,1]) / loop, sum(data_save[:,2]) / loop, sum(data_save[:,3]) / loop, sum(data_save[:,4]) / loop, sum(data_save[:,5]) / loop, sum(data_save[:,6]) / loop, sum(data_save[:,7]) / loop, sum(data_save[:,8]) / loop])

    worksheet.append(['標準差', np.std(data_save[:, 1]), np.std(data_save[:, 2]), np.std(data_save[:, 3]), np.std(data_save[:, 4]), np.std(data_save[:, 5]), np.std(data_save[:, 6]), np.std(data_save[:, 7]), np.std(data_save[:, 8])])

    worksheet.append(['數據', str(round(sum(data_save[:, 1]) / loop * 100, 2)) + ' ± ' + str(round(np.std(data_save[:, 1]) * 100, 2))
            , str(round(sum(data_save[:, 2]) / loop, 4)) + ' ± ' + str(round(np.std(data_save[:, 2]), 4))
            , str(round(sum(data_save[:, 3]) / loop, 4)) + ' ± ' + str(round(np.std(data_save[:, 3]), 4))
            , str(round(sum(data_save[:, 4]) / loop, 4)) + ' ± ' + str(round(np.std(data_save[:, 4]), 4))
            , str(round(sum(data_save[:, 5]) / loop, 4)) + ' ± ' + str(round(np.std(data_save[:, 5]), 4))
            , str(round(sum(data_save[:, 6]) / loop, 4)) + ' ± ' + str(round(np.std(data_save[:, 6]), 4))
            , str(round(sum(data_save[:, 7]) / loop, 4)) + ' ± ' + str(round(np.std(data_save[:, 7]), 4))
            , str(round(sum(data_save[:, 8]) / loop, 4)) + ' ± ' + str(round(np.std(data_save[:, 8]), 4))])

    worksheet.append([''])
    worksheet.append([''])
    worksheet.append([''])
    worksheet.append(['以下為每輪預測之結果:'])
    worksheet.append([''])

    worksheet.append([' ','Accuracy', 'Specificity', 'Sensitivity', 'Precision', "AUROC", 'AUPRC', 'F1_score', 'MCC'])

    for i in data_save :
        worksheet.append(i.tolist())

    workbook.save(excel_name)

def PR_ROC_curve(test_label, dnn_probs):
    fig = plt.gcf()
    fig.set_size_inches(16, 6)
    plt_1 = plt.subplot(121)
    plt_1.set_box_aspect(1)
    dnn_auc = roc_auc_score(test_label, dnn_probs)
    dnn_fpr, dnn_tpr, _ = roc_curve(test_label, dnn_probs)

    plt.plot(dnn_fpr, dnn_tpr, marker='.', label=' (AUROC = %0.3f)' % dnn_auc)

    plt.plot([0, 1], [0, 1], color='red', linestyle='--')  # lw 線寬
    plt.title('ROC')
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')

    plt.fill_between(dnn_fpr, dnn_tpr, color='gray', alpha=0.2)

    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.0])
    plt.legend(loc='lower right')


    plt_2 = plt.subplot(122)
    plt_2.set_box_aspect(1)
    precision, recall, thresholds = precision_recall_curve(test_label, dnn_probs)
    dnn_aup = average_precision_score(test_label, dnn_probs)

    plt.plot(recall, precision, label=' (AUPRC = %0.3f)' % dnn_aup)

    fontsize = 14
    plt.title("PR")  # 標題
    plt.xlabel("Recall", fontsize=fontsize)
    plt.ylabel("Precision", fontsize=fontsize)

    plt.fill_between(recall, precision, color='gray', alpha=0.2)


    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.0])
    plt.legend(loc='lower right')

    print("========================")
    for i in range(len(test_label)) : print(f"{test_label[i]}  {dnn_probs[i]}")

    plt.show()

# 取得時間
localtime = time.localtime()
local_time = time.strftime("%Y%m%d_%H%M%p", localtime)

###################################################   參數設定   ###########################################################

loop = 100 #幾次平均
prediction_process = 0 #每次預測結果
smote = 1

save = 1
file_name = "外部驗證結果_"
excel_name = file_name + local_time + '.xlsx'

pr_roc = 0 # 印出 PR、ROC 曲線

accuracy_indice = [3, 50, 40, 36, 83]  # 測試蛋白質編號

###################################################   導入資料   ###########################################################



SampleID = []
ProteinID = []


if pr_roc == 1 : save = 0


train_input = 'Discovery.csv'
test_input = 'Validation.csv'


excel_location = '輸出結果\\'


train_df = pd.read_csv(train_input)
test_df = pd.read_csv(test_input)

# # 查詢具空值的欄位
# display( Missing_Counts(test_df) )



###################################################   主程式   ###########################################################



train_data, train_label = data_processing(train_df)
test_data, test_label = data_processing(test_df)

goodbad = np.zeros((len(test_data)), dtype=float)

n_neighbors = [0]
leaf_size = [0]
algo = [0]
weigh = [0]
p = [0]


best_save = np.zeros((len(n_neighbors), 14), dtype=float)  # 儲存結果

train_data = train_data[:, np.array(accuracy_indice)]
test_data = test_data[:, np.array(accuracy_indice)]

result_count = 0

for n_neighbors,leaf_size,algo,weigh,p in zip(n_neighbors,leaf_size,algo,weigh,p) :

    data_save = np.zeros((loop, 9), dtype=float)  # 儲存結果


    if algo == 1:
        algorithm = 'auto'
    elif algo == 2:
        algorithm = 'ball_tree'
    elif algo == 3:
        algorithm = 'kd_tree'
    elif algo == 4:
        algorithm = 'brute'


    if weigh == 1:
        weights = 'uniform'
    elif weigh == 2:
        weights = 'distance'


    if n_neighbors == 0:
        knn = KNeighborsClassifier(n_neighbors=5)
    else:
        knn = KNeighborsClassifier(n_neighbors=n_neighbors, leaf_size=leaf_size, algorithm=algorithm, weights=weights,p=p)

    for loop_count in range(loop) :

        if smote == 1 : smote_train_data, smote_train_label = SMOTE(k_neighbors=5).fit_resample(train_data,train_label)  # 做資料平衡中的上採法 #sampling_strategy={0: }
        else : smote_train_data, smote_train_label = train_data, train_label

        knn.fit(smote_train_data, smote_train_label) #將最好的特徵組合跟訓練資料進行模型訓練

        prediction = knn.predict(test_data) #將最好的特徵組合跟測試資料進行模型預測
        prediction = prediction


        probs = knn.predict_proba(test_data)
        probs = probs[:, 1]


        if pr_roc == 1: PR_ROC_curve(test_label, probs, prediction)


        score(test_label, prediction, probs)

    if save == 1:
        sheet_name = '編號 ' + str(result_count + 1)  # 'layer ' + str(layer) + ' batch ' + str(batch_size) + ' units ' + str(units) +
        excel_save(data_save, sheet_name, excel_name, excel_location)  # 輸出結果至excel
        # print(data_save)
        print(sheet_name)

    BestSave(data_save,result_count)

    result_count = result_count + 1